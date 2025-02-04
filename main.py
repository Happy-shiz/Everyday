import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import Calendar
from datetime import datetime, timedelta
from collections import defaultdict
import calendar
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Create or connect to the Excel file
def create_excel():
    try:
        wb = openpyxl.load_workbook("calendar.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "События"
        sheet.append(["id", "title", "description", "date", "start_time", "end_time", "category"])
        wb.save("calendar.xlsx")
    return wb

wb = create_excel()
sheet = wb["События"]

# Colors for categories
CATEGORY_COLORS = {
    "Работа": "#ffcc00",
    "Семья": "#ff6699",
    "Здоровье": "#33cc33",
    "Тренировки": "#3399ff",
    "Обучение": "#9933ff",
    "События": "#ff5050"
}

# Function to add an event
def add_event():
    title = title_entry.get()
    description = description_entry.get("1.0", tk.END).strip()
    date = date_entry.get()
    start_time = start_time_entry.get()
    end_time = end_time_entry.get()
    category = category_var.get()
    if not title or not date or not start_time or not end_time or not category:
        messagebox.showwarning("Ошибка", "Заполните все обязательные поля!")
        return
    try:
        event_date = datetime.strptime(date, "%Y-%m-%d").date()
        # Check if time format is correct
        datetime.strptime(start_time, "%H:%M")
        datetime.strptime(end_time, "%H:%M")
    except ValueError:
        messagebox.showwarning("Ошибка", "Некорректный формат даты или времени!")
        return
    # Write event to Excel
    last_row = sheet.max_row + 1
    sheet.append([last_row, title, description, event_date, start_time, end_time, category])
    wb.save("calendar.xlsx")
    
    update_monthly_view()
    update_weekly_view()
    update_task_chart()
    create_gantt_chart()
    clear_inputs()

# Function to delete an event with confirmation, now linked to event labels
def delete_event(event_id):
    event_title = next((event[1] for event in events if event[0] == event_id), None)
    if event_title:
        confirm = messagebox.askyesno("Подтверждение удаления", f"Вы уверены, что хотите удалить событие '{event_title}'?")
        if confirm:
            # Search and delete event in Excel
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if row[0].value == event_id:
                    sheet.delete_rows(row[0].row)
                    wb.save("calendar.xlsx")
                    break
            update_monthly_view()
            update_weekly_view()
            update_task_chart()
            create_gantt_chart()

# Function to clear input fields
def clear_inputs():
    title_entry.delete(0, tk.END)
    description_entry.delete("1.0", tk.END)
    date_entry.delete(0, tk.END)
    start_time_entry.delete(0, tk.END)
    end_time_entry.delete(0, tk.END)

# Function to pick date
def pick_date():
    def set_date():
        date_entry.delete(0, tk.END)
        date_entry.insert(0, cal.get_date())
        top.destroy()
    top = tk.Toplevel(root)
    top.title("Выберите дату")
    cal = Calendar(top, date_pattern="yyyy-mm-dd")
    cal.pack(pady=20)
    tk.Button(top, text="Выбрать", command=set_date).pack(pady=10)

# Function to update monthly view with clickable events
def update_monthly_view():
    for widget in monthly_frame.winfo_children():
        widget.destroy()

    # Dictionary to store events by date
    global events  # Declare events as global to use it in delete_event function
    events = defaultdict(list)

    # Read events from Excel
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if len(row) == 7:  # Ensure we have the correct number of columns
            event_date, title, _, date, _, _, category = row
            if isinstance(date, datetime):
                event_date = date.date()
                events[event_date].append((row[0], f"{title} - {category}"))

    # Create header with days of the week
    days_of_week = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    day_header = '   '.join([day.ljust(4) for day in days_of_week])
    tk.Label(monthly_frame, text=day_header, font=("Courier", 12, "bold")).pack()

    # Get current year and month
    current_month = selected_month
    current_year = selected_year

    # Get the first day of the month and the number of days in the month
    first_day_of_month = datetime(current_year, current_month, 1)
    _, num_days_in_month = calendar.monthrange(current_year, current_month)

    # Create a list of days for the current month
    days_in_month = [i for i in range(1, num_days_in_month + 1)]
    
    # Format the calendar in weeks
    week = []
    for i, day in enumerate(days_in_month):
        week.append(day)
        if (i + 1) % 7 == 0 or i == len(days_in_month) - 1:
            # Print the week
            week_str = ''
            for day in week:
                day_date = datetime(current_year, current_month, day)
                events_for_day = events.get(day_date.date(), [])
                num_tasks = len(events_for_day)
                # Use ljust() to align all numbers and labels
                week_str += f"{str(day).ljust(3)}({str(num_tasks).rjust(2)})   "
            tk.Label(monthly_frame, text=week_str, font=("Courier", 10)).pack()
            week = []

# Function to update weekly view
def update_weekly_view():
    for widget in weekly_frame.winfo_children():
        widget.destroy()

    try:
        selected_date = datetime.strptime(date_entry.get(), "%Y-%m-%d").date()
    except ValueError:
        selected_date = datetime.now().date()
    
    start_of_week = selected_date - timedelta(days=selected_date.weekday())
    days = [start_of_week + timedelta(days=i) for i in range(7)]

    # Create headers
    headers = [day.strftime("%a") for day in days]
    for i, header in enumerate(headers):
        tk.Label(weekly_frame, text=header, font=("Arial", 10, "bold"), borderwidth=1, relief="solid").grid(row=0, column=i+1, sticky="nsew")

    # Time labels on the left
    times = [f"{h:02d}:00" for h in range(24)]
    for i, time in enumerate(times):
        tk.Label(weekly_frame, text=time, font=("Arial", 8)).grid(row=i+1, column=0, sticky="nsew")

    # Fill in the days and events
    for i, day in enumerate(days):
        events_for_day = events.get(day, [])
        for event in events_for_day:
            # Event is now a tuple like (event_id, event_string)
            event_id, event_string = event
            # Parse the event string to get time and title
            title, rest = event_string.split(' - ', 1)
            try:
                time_str = rest.split(',')[0].strip()
                if ':' in time_str:
                    event_time = datetime.strptime(time_str, "%H:%M").time().hour
                elif time_str.isdigit():  # Check if it's just an hour
                    event_time = int(time_str)
                else:
                    continue  # Skip if no valid time found
            except ValueError:
                print(f"Could not parse time for event: {event_string}")
                continue  # Skip this event if time can't be parsed
            
            tk.Label(weekly_frame, text=title, bg=CATEGORY_COLORS.get(rest.split(',')[1].strip(), "grey"), borderwidth=1, relief="solid").grid(row=event_time+1, column=i+1, sticky="nsew")

# Function to update task chart
def update_task_chart():
    for widget in chart_frame.winfo_children():
        widget.destroy()

    try:
        selected_date = datetime.strptime(date_entry.get(), "%Y-%m-%d").date()
    except ValueError:
        selected_date = datetime.now().date()
    
    start_date = selected_date
    for i in range(7):
        day = start_date + timedelta(days=i)
        events_count = len(events.get(day, []))
        bar_length = min(events_count * 20, 200)  # Limit bar length for visualization
        
        # Draw bar
        bar = tk.Canvas(chart_frame, width=bar_length, height=20, bg=CATEGORY_COLORS.get("Работа", "grey"))
        bar.create_text(10, 10, text=f"{day.strftime('%d %b')} ({events_count})", anchor="w")
        bar.grid(row=i, column=0, sticky="w")
    
    # Add labels for days
    for i in range(7):
        day_label = tk.Label(chart_frame, text=start_date.strftime("%d %b"), borderwidth=1, relief="solid")
        day_label.grid(row=i, column=1, sticky="w")
        start_date += timedelta(days=1)

# Function to create Gantt chart with week switching
current_week = [datetime.now().date() - timedelta(days=datetime.now().weekday())]

def create_gantt_chart():
    for widget in gantt_frame.winfo_children():
        widget.destroy()

    fig, ax = plt.subplots(figsize=(10, 6))

    # Read tasks from Excel for the current week
    tasks = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if len(row) == 7:
            id, title, description, date, start_time, end_time, category = row
            if isinstance(date, datetime) and current_week[0] <= date.date() <= current_week[0] + timedelta(days=6):
                # Handle None or empty string values for time
                start_time = start_time or "00:00"  # Default to midnight if start_time is None
                end_time = end_time or "23:59"  # Default to end of day if end_time is None
                
                try:
                    tasks.append({
                        'title': title,
                        'date': date,
                        'start_time': datetime.strptime(start_time, "%H:%M").time(),
                        'end_time': datetime.strptime(end_time, "%H:%M").time(),
                        'category': category
                    })
                except ValueError:
                    # If parsing fails, skip this task or log the error
                    print(f"Failed to parse time for task: {title}")

    # Sort tasks by date and start time
    tasks.sort(key=lambda x: (x['date'], x['start_time']))

    # Set up the plot
    start_date = current_week[0]
    end_date = current_week[0] + timedelta(days=6)
    date_range = end_date - start_date
    days = date_range.days + 1

    ax.set_xlim(0, days)
    ax.set_ylim(0, len(tasks))
    ax.set_yticks(range(len(tasks)))
    ax.set_yticklabels([f"{task['start_time'].strftime('%H:%M')} - {task['end_time'].strftime('%H:%M')}" for task in tasks])
    ax.set_xticks(range(days))
    ax.set_xticklabels([(start_date + timedelta(days=i)).strftime("%a") for i in range(days)])
    ax.invert_yaxis()

    # Plot tasks
    for i, task in enumerate(tasks):
        if isinstance(task['date'], datetime):
            day = (task['date'].date() - start_date).days
            start_hour = task['start_time'].hour + task['start_time'].minute / 60
            end_hour = task['end_time'].hour + task['end_time'].minute / 60
            duration = end_hour - start_hour
            ax.barh(i, duration, left=day + start_hour / 24, height=0.5, color=CATEGORY_COLORS.get(task['category'], "grey"))
            # Center text on the bar
            center = day + (start_hour + end_hour) / 2 / 24
            ax.text(center, i, task['title'], ha='center', va='center', fontsize=8)

    ax.set_xlabel('День')
    ax.set_ylabel('Время')
    ax.set_title('График Гантта')
    ax.grid(True)

    # Embed the plot in Tkinter
    canvas = FigureCanvasTkAgg(fig, master=gantt_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # Week navigation buttons
    tk.Button(gantt_frame, text="Предыдущая неделя", command=prev_week).pack(side=tk.LEFT, padx=5)
    tk.Button(gantt_frame, text="Следующая неделя", command=next_week).pack(side=tk.LEFT, padx=5)

def next_week():
    current_week[0] += timedelta(days=7)
    create_gantt_chart()

def prev_week():
    current_week[0] -= timedelta(days=7)
    create_gantt_chart()

# Function to switch to the previous year
def prev_year():
    global selected_year
    selected_year -= 1
    year_label.config(text=f"Год: {selected_year}")
    update_monthly_view()
    update_weekly_view()
    update_task_chart()
    create_gantt_chart()

# Function to switch to the next year
def next_year():
    global selected_year
    selected_year += 1
    year_label.config(text=f"Год: {selected_year}")
    update_monthly_view()
    update_weekly_view()
    update_task_chart()
    create_gantt_chart()

# Function to switch to the previous month
def prev_month():
    global selected_month
    if selected_month == 1:
        selected_month = 12
        prev_year()
    else:
        selected_month -= 1
    month_label.config(text=f"{calendar.month_name[selected_month]} {selected_year}")
    update_monthly_view()
    update_weekly_view()
    update_task_chart()
    create_gantt_chart()

# Function to switch to the next month
def next_month():
    global selected_month
    if selected_month == 12:
        selected_month = 1
        next_year()
    else:
        selected_month += 1
    month_label.config(text=f"{calendar.month_name[selected_month]} {selected_year}")
    update_monthly_view()
    update_weekly_view()
    update_task_chart()
    create_gantt_chart()

# Create GUI
root = tk.Tk()
root.title("Офлайн Календарь")

# Get screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Set window size to about 90% of screen size
window_width = int(screen_width * 0.9)
window_height = int(screen_height * 0.9)
root.geometry(f"{window_width}x{window_height}")

# Notebook for tabs
notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill="both", padx=10, pady=10)

# Monthly view frame
monthly_frame = tk.Frame(notebook)
notebook.add(monthly_frame, text="Месяц")

# Weekly view frame
weekly_frame = tk.Frame(notebook)
notebook.add(weekly_frame, text="Неделя")

# Chart frame
chart_frame = tk.Frame(notebook)
notebook.add(chart_frame, text="График задач")

# Gantt chart frame
gantt_frame = tk.Frame(notebook)
notebook.add(gantt_frame, text="График Гантта")

# Set initial year and month
selected_year = datetime.now().year
selected_month = datetime.now().month

# Navigation buttons for year and month
year_nav_frame = tk.Frame(root)
year_nav_frame.pack(pady=10)

prev_year_btn = tk.Button(year_nav_frame, text="<<", command=prev_year)
prev_year_btn.pack(side="left", padx=5)
year_label = tk.Label(year_nav_frame, text=f"Год: {selected_year}", font=("Arial", 12, "bold"))
year_label.pack(side="left")
next_year_btn = tk.Button(year_nav_frame, text=">>", command=next_year)
next_year_btn.pack(side="left", padx=5)

month_nav_frame = tk.Frame(root)
month_nav_frame.pack(pady=5)

prev_month_btn = tk.Button(month_nav_frame, text="<", command=prev_month)
prev_month_btn.pack(side="left", padx=5)

# Fixed width for month label to prevent jumping
month_label = tk.Label(month_nav_frame, text=f"{calendar.month_name[selected_month]} {selected_year}", font=("Arial", 12, "bold"), width=15, anchor="w")
month_label.pack(side="left")

next_month_btn = tk.Button(month_nav_frame, text=">", command=next_month)
next_month_btn.pack(side="left", padx=5)

# Input fields with labels
input_frame = tk.Frame(root)
input_frame.pack(pady=10)

tk.Label(input_frame, text="Название события:").grid(row=0, column=0, sticky="w")
title_entry = tk.Entry(input_frame, width=50)
title_entry.grid(row=0, column=1, sticky="w")

tk.Label(input_frame, text="Описание события:").grid(row=1, column=0, sticky="w")
description_entry = tk.Text(input_frame, height=3, width=50)
description_entry.grid(row=1, column=1, sticky="w")

tk.Label(input_frame, text="Дата события:").grid(row=2, column=0, sticky="w")
date_entry = tk.Entry(input_frame, width=50)
date_entry.grid(row=2, column=1, sticky="w")
tk.Button(input_frame, text="Выбрать дату", command=pick_date).grid(row=2, column=2, sticky="w")

tk.Label(input_frame, text="Время начала:").grid(row=3, column=0, sticky="w")
start_time_entry = tk.Entry(input_frame, width=25)
start_time_entry.grid(row=3, column=1, sticky="w")

tk.Label(input_frame, text="Время окончания:").grid(row=3, column=2, sticky="w")
end_time_entry = tk.Entry(input_frame, width=25)
end_time_entry.grid(row=3, column=3, sticky="w")

tk.Label(input_frame, text="Категория события:").grid(row=4, column=0, sticky="w")
category_var = tk.StringVar()
category_dropdown = ttk.Combobox(input_frame, textvariable=category_var, values=list(CATEGORY_COLORS.keys()), width=48)
category_dropdown.grid(row=4, column=1, sticky="w")

tk.Button(input_frame, text="Добавить событие", command=add_event).grid(row=5, column=1, pady=5)

# Button to refresh the views after deletions
tk.Button(root, text="Обновить календарь", command=lambda: [update_monthly_view(), update_weekly_view(), update_task_chart(), create_gantt_chart()]).pack()

update_monthly_view()
update_weekly_view()
update_task_chart()
create_gantt_chart()

root.mainloop()
wb.save("calendar.xlsx")